from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from app.core.config import settings
from app.services.admin_dashboard_service import AdminDashboardService
from app.services.attendance_service import AttendanceService
from app.services.bonus_service import BonusService
from app.services.db_service import DatabaseService
from app.services.invoice_service import InvoiceService
from app.services.sheets_service import GoogleSheetsService
from app.services.signup_service import SignupService


@dataclass
class ExportService:
    db: DatabaseService
    sheets: GoogleSheetsService
    signup_service: SignupService
    attendance_service: AttendanceService
    invoice_service: InvoiceService
    bonus_service: BonusService
    dashboard_service: AdminDashboardService

    def export_google_sheet(self, target_date: str) -> dict:
        if not self.sheets.is_available():
            return {'ok': False, 'message': 'Google Sheets 尚未設定完成'}
        signup_count = self.signup_service.replace_signup_sheet(target_date)
        attendance_count = self.attendance_service.replace_attendance_sheet(target_date)
        attendance_rows = [r for r in self.db.list_all_attendance() if r['date'] == target_date]
        invoice_rows = [
            self.invoice_service.build_invoice_row(
                uid=r['uid'], name=r['name'], date_key=r['date'], shift=r.get('shift', '') or '', in_time=r.get('in_time', '') or '',
                out_time=r.get('out_time', '') or '', break_start=r.get('break_start', '') or '', break_end=r.get('break_end', '') or '',
                break_min=r.get('break_min'), net_min=r.get('net_min'), sync_status='已同步' if r.get('out_time') else '待同步'
            )
            for r in attendance_rows
        ]
        self.invoice_service.replace_invoice_sheet(invoice_rows)
        self.db.add_log('INFO', 'export.google_sheet', f'date={target_date} signup={signup_count} attendance={attendance_count} invoice={len(invoice_rows)}')
        return {'ok': True, 'message': f'已匯出 {target_date} 資料到 Google Sheet', 'url': f'https://docs.google.com/spreadsheets/d/{settings.google_spreadsheet_id}/edit' if settings.google_spreadsheet_id else ''}

    def export_excel(self, target_date: str) -> Path:
        payload = self.dashboard_service.get_dashboard_payload(target_date)
        wb = Workbook()
        ws1 = wb.active
        ws1.title = settings.signup_sheet_name
        ws1.append(['UID', '姓名', '群組ID', '班別', '報班窗口', '登記時間', '狀態'])
        for row in payload['signups']:
            ws1.append([row.get('uid',''), row.get('name',''), row.get('groupId',''), row.get('shift',''), row.get('window',''), row.get('signupAt',''), row.get('status','')])

        ws2 = wb.create_sheet(settings.att_sheet_name)
        ws2.append(['日期', 'UID', '姓名', '群組ID', '上班時間', '休息開始', '休息結束', '下班時間', '休息分鐘', '工作分鐘', '淨工作分鐘', '備註', '狀態'])
        for row in payload['attendance']:
            ws2.append([row.get('date',''), row.get('uid',''), row.get('name',''), row.get('groupId',''), row.get('in',''), row.get('breakStart',''), row.get('breakEnd',''), row.get('out',''), row.get('breakMin',''), row.get('workMin',''), row.get('netMin',''), row.get('remark',''), row.get('status','')])

        ws3 = wb.create_sheet(settings.invoice_sheet_name)
        ws3.append(['UID', '日期', '班別', '姓名', '上班時間', '下班時間', '出勤總時數', '休息開始', '休息結束', '休息時間', '實際出勤總時數', '時薪', '獎金(每小時)', '薪資', '獎金總額', '同步狀態'])
        for row in payload['invoices']:
            ws3.append([row.get('uid',''), row.get('date',''), row.get('shift',''), row.get('name',''), row.get('in',''), row.get('out',''), row.get('totalHours',''), row.get('breakStart',''), row.get('breakEnd',''), row.get('breakHours',''), row.get('netHours',''), row.get('wage',''), row.get('bonus',''), row.get('salary',''), row.get('bonusAmount',''), row.get('sync','')])

        ws4 = wb.create_sheet(settings.bonus_sheet_name)
        ws4.append(['日期', '來源班別', '對應班別別名', '獎金(每小時)', '來源分頁', '狀態'])
        for row in payload['bonusRules']:
            ws4.append([row.get('date',''), row.get('sourceShift',''), row.get('alias',''), row.get('bonusPerHour',''), row.get('sourceSheet',''), row.get('status','')])

        path = Path(settings.export_dir) / f'line_attendance_export_{target_date}_{datetime.now().strftime("%H%M%S")}.xlsx'
        wb.save(path)
        self.db.add_log('INFO', 'export.excel', f'date={target_date} file={path.name}')
        return path
